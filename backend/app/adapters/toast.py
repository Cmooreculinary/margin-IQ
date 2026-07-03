"""Toast PMIX export adapter (CSV/XLSX) -- the v1 first-class citizen.

Expected Toast PMIX export columns (header names are matched case-insensitively
and with underscores/spaces interchangeable):
    PLU / Menu Item Number
    Menu Item / Item Name
    Item Qty / Qty Sold
    Net Amount / Net Sales
    Sales Date range is passed in explicitly since Toast exports don't always
    carry it per-row.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import BinaryIO

from openpyxl import load_workbook

from app.adapters.base import PosAdapter

_HEADER_ALIASES = {
    "plu": {"plu", "menu item number", "item number", "sku"},
    "item_name": {"menu item", "item name", "item"},
    "units_sold": {"item qty", "qty sold", "qty", "quantity"},
    "gross_revenue": {"net amount", "net sales", "gross amount", "sales"},
}


def _normalize_header(h: str) -> str:
    return h.strip().lower().replace("_", " ")


def _map_headers(headers: list[str]) -> dict[str, int]:
    normalized = [_normalize_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for idx, h in enumerate(normalized):
            if h in aliases:
                mapping[field] = idx
                break
        if field not in mapping:
            raise ValueError(f"Toast PMIX export missing a recognizable column for '{field}'")
    return mapping


class ToastAdapter(PosAdapter):
    name = "toast"

    def parse_pmix(
        self,
        file: BinaryIO,
        location_id: str,
        period_start: datetime | None = None,
        period_end: datetime | None = None,
    ) -> list[dict]:
        filename = getattr(file, "name", "") or ""
        raw = file.read()
        if filename.lower().endswith(".xlsx"):
            rows = self._parse_xlsx(raw)
        else:
            rows = self._parse_csv(raw)

        if not rows:
            return []

        mapping = _map_headers(rows[0])
        out = []
        for row in rows[1:]:
            if not any(row):
                continue
            plu = str(row[mapping["plu"]]).strip()
            if not plu:
                continue
            try:
                units_sold = int(float(row[mapping["units_sold"]]))
                gross_revenue = float(
                    str(row[mapping["gross_revenue"]]).replace("$", "").replace(",", "")
                )
            except (ValueError, TypeError):
                continue
            out.append(
                {
                    "location_id": location_id,
                    "plu": plu,
                    "item_name": str(row[mapping["item_name"]]).strip(),
                    "period_start": period_start,
                    "period_end": period_end,
                    "units_sold": units_sold,
                    "gross_revenue": gross_revenue,
                    "source": self.name,
                }
            )
        return out

    @staticmethod
    def _parse_csv(raw: bytes) -> list[list[str]]:
        text = raw.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader]

    @staticmethod
    def _parse_xlsx(raw: bytes) -> list[list]:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
