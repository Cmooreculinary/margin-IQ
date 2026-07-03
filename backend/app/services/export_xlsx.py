"""XLSX exports: item analysis table and the approved-action implementation
checklist. Every export is watermarked with tenant + timestamp per the
tenant-isolation guardrail."""
from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FIRE_ORANGE = "EC5B13"
OBSIDIAN = "0D0D0D"


def _watermark(ws, tenant_name: str):
    ws.append([])
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    footer = ws.cell(row=ws.max_row + 1, column=1, value=f"MARGIN IQ | {tenant_name} | Exported {ts}")
    footer.font = Font(italic=True, size=9, color="888888")


def build_item_analysis_workbook(*, tenant_name: str, items: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Analysis"

    headers = [
        "PLU", "Item", "Category", "Location", "Units Sold", "Price",
        "Food Cost", "Labor Cost", "Packaging", "Prime Cost %", "CM $", "CM %",
        "Quadrant", "Food-Cost Mirage",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color=FIRE_ORANGE, end_color=FIRE_ORANGE, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for item in items:
        ws.append(
            [
                item.get("plu"),
                item.get("name"),
                item.get("category"),
                item.get("location_name"),
                item.get("units_sold"),
                item.get("price"),
                item.get("food_cost"),
                item.get("labor_cost"),
                item.get("packaging_cost"),
                item.get("prime_cost_pct"),
                item.get("cm_dollars"),
                item.get("cm_pct"),
                item.get("quadrant"),
                "YES" if item.get("is_food_cost_mirage") else "",
            ]
        )

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    _watermark(ws, tenant_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_implementation_checklist_workbook(*, tenant_name: str, approved_recs: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Implementation Checklist"

    headers = [
        "Location", "PLU", "Item", "Move Type", "Current Price", "Final Price",
        "Rationale", "Projected BPS Lift", "Approved By", "Approved At",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color=FIRE_ORANGE, end_color=FIRE_ORANGE, fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for rec in approved_recs:
        ws.append(
            [
                rec.get("location_name"),
                rec.get("plu"),
                rec.get("name"),
                rec.get("type"),
                rec.get("current_price"),
                rec.get("final_price", rec.get("recommended_price")),
                rec.get("rationale"),
                rec.get("projected_bps_lift"),
                rec.get("decided_by"),
                rec.get("decided_at"),
            ]
        )

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    _watermark(ws, tenant_name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
